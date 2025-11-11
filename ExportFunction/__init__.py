"""
ExportFunction - Gera arquivo Excel da promoção e retorna como base64
"""
import logging
import json
import azure.functions as func
from datetime import datetime
from io import BytesIO
import base64

logger = logging.getLogger(__name__)

# Import do openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
    logger.info("✅ openpyxl carregado com sucesso")
except ImportError as e:
    EXCEL_AVAILABLE = False
    logger.error(f"❌ openpyxl não disponível: {e}")

async def main(req: func.HttpRequest) -> func.HttpResponse:
    """
    Gera arquivo Excel da promoção e retorna como base64
    
    POST /api/export
    Body: { "promo_data": {...}, "format": "excel" }
    """
    logger.info('📊 ExportFunction: Gerando exportação')
    
    try:
        req_body = req.get_json()
        promo_data = req_body.get('promo_data', {})
        export_format = req_body.get('format', 'excel')
        
        if not promo_data:
            return func.HttpResponse(
                json.dumps({"success": False, "error": "promo_data é obrigatório"}),
                mimetype="application/json",
                status_code=400
            )
        
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl não está disponível")
            return func.HttpResponse(
                json.dumps({"success": False, "error": "openpyxl não disponível"}),
                mimetype="application/json",
                status_code=500
            )
        
        # Gera arquivo Excel
        logger.info(f"📝 Gerando arquivo Excel para: {promo_data.get('titulo', 'sem_titulo')}")
        try:
            excel_buffer = generate_excel(promo_data)
            logger.info("✅ Excel gerado em memória")
        except Exception as e:
            logger.error(f"❌ Erro ao gerar Excel: {str(e)}", exc_info=True)
            return func.HttpResponse(
                json.dumps({"success": False, "error": f"Erro ao gerar Excel: {str(e)}"}),
                mimetype="application/json",
                status_code=500
            )
        
        # Converte para base64
        try:
            excel_buffer.seek(0)
            excel_bytes = excel_buffer.read()
            logger.info(f"📦 Tamanho do Excel: {len(excel_bytes)} bytes")
            excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
            logger.info(f"📦 Base64 gerado: {len(excel_base64)} caracteres")
        except Exception as e:
            logger.error(f"❌ Erro ao converter para base64: {str(e)}", exc_info=True)
            return func.HttpResponse(
                json.dumps({"success": False, "error": f"Erro ao converter: {str(e)}"}),
                mimetype="application/json",
                status_code=500
            )
        
        # Nome do arquivo
        try:
            titulo = promo_data.get('titulo', 'promocao')
            # Remove caracteres especiais
            titulo = ''.join(c for c in titulo if c.isalnum() or c in (' ', '_')).replace(' ', '_').lower()
            timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            filename = f"{titulo}_{timestamp}.xlsx"
            logger.info(f"✅ Exportação concluída: {filename}")
        except Exception as e:
            logger.warning(f"Erro ao gerar filename: {e}")
            filename = f"promocao_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return func.HttpResponse(
            json.dumps({
                "success": True,
                "excel_base64": excel_base64,
                "filename": filename,
                "format": "excel"
            }, ensure_ascii=False),
            mimetype="application/json",
            status_code=200
        )
        
    except Exception as e:
        logger.error(f"❌ Erro ao gerar exportação: {str(e)}", exc_info=True)
        return func.HttpResponse(
            json.dumps({"success": False, "error": str(e)}),
            mimetype="application/json",
            status_code=500
        )


def generate_excel(promo_data: dict) -> BytesIO:
    """
    Gera arquivo Excel com os dados da promoção em formato de tabela
    Headers nas colunas, dados nas linhas
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Promoções"
    
    # Estilos
    header_fill = PatternFill(start_color="1F3C88", end_color="1F3C88", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Define colunas (headers)
    headers = [
        "Título",
        "Mecânica", 
        "Descrição",
        "Segmentação",
        "Canal",
        "Categoria",
        "Período Início",
        "Período Fim",
        "Condições",
        "Recompensas",
        "Produtos",
        "Desconto %",
        "Quantidade Mínima"
    ]
    
    # Mapeamento header -> campo
    field_mapping = {
        "Título": "titulo",
        "Mecânica": "mecanica",
        "Descrição": "descricao",
        "Segmentação": "segmentacao",
        "Canal": "canal",
        "Categoria": "categoria",
        "Período Início": "periodo_inicio",
        "Período Fim": "periodo_fim",
        "Condições": "condicoes",
        "Recompensas": "recompensas",
        "Produtos": "produtos",
        "Desconto %": "desconto_percentual",
        "Quantidade Mínima": "qt_minima"
    }
    
    # Escreve headers na primeira linha
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Verifica se é múltiplas promoções ou uma única
    promos_to_write = []
    
    if "multiple_promotions" in promo_data and promo_data["multiple_promotions"]:
        # Múltiplas promoções
        promos_to_write = promo_data["multiple_promotions"]
        logger.info(f"📊 Gerando Excel com {len(promos_to_write)} promoções")
    else:
        # Uma única promoção
        promos_to_write = [promo_data]
        logger.info("📊 Gerando Excel com 1 promoção")
    
    # Escreve dados de cada promoção
    for row_idx, promo in enumerate(promos_to_write, start=2):
        for col_idx, header in enumerate(headers, start=1):
            field_name = field_mapping[header]
            value = promo.get(field_name, "")
            
            # Trata valores
            if value is None:
                value = ""
            elif isinstance(value, list):
                value = ", ".join(str(item) for item in value if item)
            else:
                value = str(value) if value else ""
            
            # Escreve valor
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
    
    # Ajusta largura das colunas
    column_widths = {
        1: 25,   # Título
        2: 15,   # Mecânica
        3: 40,   # Descrição
        4: 30,   # Segmentação
        5: 12,   # Canal
        6: 18,   # Categoria
        7: 15,   # Período Início
        8: 15,   # Período Fim
        9: 35,   # Condições
        10: 25,  # Recompensas
        11: 30,  # Produtos
        12: 12,  # Desconto %
        13: 18   # Qt Mínima
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Altura da linha de header
    ws.row_dimensions[1].height = 30
    
    # Congela primeira linha (headers)
    ws.freeze_panes = "A2"
    
    # Salva em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
