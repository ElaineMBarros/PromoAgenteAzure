"""
Blob Storage Adapter - Substitui armazenamento local por Azure Blob Storage
Gerencia upload/download de arquivos Excel gerados
Compatível com Python 3.11 e Azure Functions
"""
import os
import logging
from typing import Optional
from datetime import datetime
from pathlib import Path
from io import BytesIO
from azure.storage.blob import BlobServiceClient, ContentSettings

logger = logging.getLogger(__name__)


class BlobStorageAdapter:
    """Adapter para Azure Blob Storage - substitui armazenamento local"""
    
    def __init__(self):
        """Inicializa conexão com Blob Storage usando variáveis de ambiente"""
        self.connection_string = os.environ.get("AZURE_STORAGE_CONNECTION_STRING")
        self.container_name = "excel-exports"
        
        if not self.connection_string:
            logger.warning("⚠️ Blob Storage não configurado. Arquivos serão salvos localmente.")
            self.blob_service_client = None
            self.container_client = None
            # Fallback para armazenamento local
            self.local_storage_path = Path("exports")
            self.local_storage_path.mkdir(exist_ok=True)
            return
        
        try:
            self.blob_service_client = BlobServiceClient.from_connection_string(
                self.connection_string
            )
            
            # Cria ou obtém container
            self.container_client = self.blob_service_client.get_container_client(
                self.container_name
            )
            
            # Tenta criar o container se não existir
            try:
                self.container_client.create_container()
                logger.info(f"Container '{self.container_name}' criado")
            except Exception:
                # Container já existe
                pass
            
            logger.info("✅ Blob Storage conectado com sucesso!")
            
        except Exception as e:
            logger.error(f"❌ Erro ao conectar Blob Storage: {e}")
            self.blob_service_client = None
            self.container_client = None
            self.local_storage_path = Path("exports")
            self.local_storage_path.mkdir(exist_ok=True)
    
    def upload_excel_file(
        self, 
        file_content: bytes, 
        filename: str,
        promo_id: str = None
    ) -> str:
        """
        Upload de arquivo Excel para Blob Storage
        
        Args:
            file_content: Conteúdo do arquivo em bytes
            filename: Nome do arquivo
            promo_id: ID da promoção (opcional)
            
        Returns:
            str: URL do arquivo ou caminho local
        """
        # Se Blob Storage não está configurado, salva localmente
        if not self.container_client:
            return self._save_local(file_content, filename)
        
        try:
            # Organiza arquivos por data: 2025/11/06/arquivo.xlsx
            date_path = datetime.utcnow().strftime("%Y/%m/%d")
            blob_name = f"{date_path}/{filename}"
            
            # Upload do arquivo
            blob_client = self.container_client.get_blob_client(blob_name)
            
            # Define content type para Excel
            content_settings = ContentSettings(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            blob_client.upload_blob(
                file_content,
                overwrite=True,
                content_settings=content_settings
            )
            
            # Retorna URL do blob
            blob_url = blob_client.url
            logger.info(f"✅ Arquivo enviado para Blob Storage: {blob_name}")
            
            return blob_url
            
        except Exception as e:
            logger.error(f"❌ Erro ao enviar para Blob Storage: {e}")
            # Fallback para local
            return self._save_local(file_content, filename)
    
    def upload_excel_from_path(self, file_path: str) -> str:
        """
        Upload de arquivo Excel a partir de caminho local
        
        Args:
            file_path: Caminho do arquivo local
            
        Returns:
            str: URL do arquivo ou caminho local
        """
        try:
            with open(file_path, 'rb') as f:
                file_content = f.read()
            
            filename = Path(file_path).name
            return self.upload_excel_file(file_content, filename)
            
        except Exception as e:
            logger.error(f"Erro ao fazer upload do arquivo {file_path}: {e}")
            return file_path
    
    def download_excel_file(self, blob_url: str) -> Optional[bytes]:
        """
        Download de arquivo Excel do Blob Storage
        
        Args:
            blob_url: URL do blob
            
        Returns:
            bytes: Conteúdo do arquivo ou None
        """
        if not self.container_client:
            logger.warning("Blob Storage não configurado")
            return None
        
        try:
            # Extrai nome do blob da URL
            blob_name = blob_url.split(f"{self.container_name}/")[-1]
            
            blob_client = self.container_client.get_blob_client(blob_name)
            download_stream = blob_client.download_blob()
            
            return download_stream.readall()
            
        except Exception as e:
            logger.error(f"Erro ao fazer download: {e}")
            return None
    
    def get_blob_url(self, filename: str) -> Optional[str]:
        """
        Obtém URL de um arquivo no Blob Storage
        
        Args:
            filename: Nome do arquivo
            
        Returns:
            str: URL do blob ou None
        """
        if not self.container_client:
            return None
        
        try:
            # Organiza por data
            date_path = datetime.utcnow().strftime("%Y/%m/%d")
            blob_name = f"{date_path}/{filename}"
            
            blob_client = self.container_client.get_blob_client(blob_name)
            return blob_client.url
            
        except Exception as e:
            logger.error(f"Erro ao obter URL: {e}")
            return None
    
    def list_excel_files(self, prefix: str = None, limit: int = 50) -> list:
        """
        Lista arquivos Excel no Blob Storage
        
        Args:
            prefix: Prefixo para filtrar (ex: "2025/11/06/")
            limit: Número máximo de arquivos
            
        Returns:
            list: Lista de nomes de blobs
        """
        if not self.container_client:
            # Lista arquivos locais
            return self._list_local_files()
        
        try:
            blobs = self.container_client.list_blobs(name_starts_with=prefix)
            
            files = []
            for i, blob in enumerate(blobs):
                if i >= limit:
                    break
                files.append({
                    'name': blob.name,
                    'size': blob.size,
                    'created': blob.creation_time,
                    'url': f"{self.container_client.url}/{blob.name}"
                })
            
            return files
            
        except Exception as e:
            logger.error(f"Erro ao listar arquivos: {e}")
            return []
    
    def delete_excel_file(self, blob_name: str) -> bool:
        """
        Deleta um arquivo do Blob Storage
        
        Args:
            blob_name: Nome do blob
            
        Returns:
            bool: True se deletou com sucesso
        """
        if not self.container_client:
            return False
        
        try:
            blob_client = self.container_client.get_blob_client(blob_name)
            blob_client.delete_blob()
            logger.info(f"Arquivo deletado: {blob_name}")
            return True
            
        except Exception as e:
            logger.error(f"Erro ao deletar arquivo: {e}")
            return False
    
    # ========== MÉTODOS INTERNOS - FALLBACK LOCAL ==========
    
    def _save_local(self, file_content: bytes, filename: str) -> str:
        """Salva arquivo localmente como fallback"""
        try:
            filepath = self.local_storage_path / filename
            with open(filepath, 'wb') as f:
                f.write(file_content)
            
            logger.info(f"📁 Arquivo salvo localmente: {filepath}")
            return str(filepath.absolute())
            
        except Exception as e:
            logger.error(f"Erro ao salvar localmente: {e}")
            raise
    
    def _list_local_files(self) -> list:
        """Lista arquivos locais"""
        try:
            files = []
            for filepath in self.local_storage_path.glob("*.xlsx"):
                files.append({
                    'name': filepath.name,
                    'size': filepath.stat().st_size,
                    'created': datetime.fromtimestamp(filepath.stat().st_ctime),
                    'url': str(filepath.absolute())
                })
            return files
            
        except Exception as e:
            logger.error(f"Erro ao listar arquivos locais: {e}")
            return []


class ExcelServiceAzure:
    """
    Serviço de Excel integrado com Blob Storage
    Mantém compatibilidade com código existente
    """
    
    def __init__(self):
        self.blob_adapter = BlobStorageAdapter()
    
    def generate_promotion_excel(self, promo_data: dict) -> str:
        """
        Gera Excel e faz upload para Blob Storage
        
        Args:
            promo_data: Dados da promoção
            
        Returns:
            str: URL do arquivo no Blob ou caminho local
        """
        # Importa o serviço original
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        from io import BytesIO
        
        # Cria workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Promoção"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers e dados (versão simplificada)
        headers = ["Título", "Mecânica", "Descrição", "Público", "Início", "Fim"]
        ws.append(headers)
        
        # Estiliza cabeçalhos
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Dados
        data_row = [
            promo_data.get('titulo', ''),
            promo_data.get('mecanica', ''),
            promo_data.get('descricao', ''),
            promo_data.get('segmentacao', ''),
            promo_data.get('periodo_inicio', ''),
            promo_data.get('periodo_fim', '')
        ]
        ws.append(data_row)
        
        # Salva em memória
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Nome do arquivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        titulo_safe = promo_data.get('titulo', 'promocao').replace(' ', '_')[:30]
        filename = f"promocao_{titulo_safe}_{timestamp}.xlsx"
        
        # Upload para Blob Storage
        blob_url = self.blob_adapter.upload_excel_file(
            excel_buffer.read(),
            filename,
            promo_data.get('promo_id')
        )
        
        return blob_url


# Instâncias globais
blob_adapter = BlobStorageAdapter()
excel_service_azure = ExcelServiceAzure()
