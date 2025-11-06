"""
Excel Service - Gera arquivos Excel com dados das promoções
"""
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from typing import Dict

logger = logging.getLogger(__name__)


class ExcelService:
    """Serviço para gerar arquivos Excel com dados de promoções"""
    
    def __init__(self, output_dir: str = "exports"):
        """
        Inicializa o serviço de Excel
        
        Args:
            output_dir: Diretório onde os arquivos serão salvos
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        logger.info(f"📊 ExcelService inicializado - Diretório: {self.output_dir}")
    
    def _dividir_por_meses(self, promo_data: Dict) -> list:
        """
        Divide uma promoção em múltiplas promoções mensais se o período > 1 mês
        
        Args:
            promo_data: Dados da promoção original
            
        Returns:
            list: Lista de promoções (uma por mês)
        """
        from datetime import datetime
        from dateutil.relativedelta import relativedelta
        
        periodo_inicio = promo_data.get('periodo_inicio', '')
        periodo_fim = promo_data.get('periodo_fim', '')
        
        if not periodo_inicio or not periodo_fim:
            return [promo_data]
        
        try:
            # Parseia as datas
            inicio = datetime.strptime(periodo_inicio, '%d/%m/%Y')
            fim = datetime.strptime(periodo_fim, '%d/%m/%Y')
            
            # Calcula diferença em meses
            meses_diff = (fim.year - inicio.year) * 12 + (fim.month - inicio.month)
            
            # Se for mesmo mês ou menos de 1 mês completo, retorna original
            if meses_diff < 1:
                return [promo_data]
            
            # Divide em promoções mensais
            promocoes_mensais = []
            data_atual = inicio
            titulo_base = promo_data.get('titulo', 'Promoção')
            
            while data_atual <= fim:
                # Calcula início e fim do mês atual
                primeiro_dia = data_atual.replace(day=1)
                
                # Último dia do mês
                proximo_mes = primeiro_dia + relativedelta(months=1)
                ultimo_dia = proximo_mes - relativedelta(days=1)
                
                # Ajusta para não ultrapassar a data final
                if ultimo_dia > fim:
                    ultimo_dia = fim
                
                # Cria cópia da promoção para este mês
                promo_mes = promo_data.copy()
                promo_mes['periodo_inicio'] = primeiro_dia.strftime('%d/%m/%Y')
                promo_mes['periodo_fim'] = ultimo_dia.strftime('%d/%m/%Y')
                
                # Atualiza título com o mês
                mes_nome = primeiro_dia.strftime('%B %Y')
                meses_pt = {
                    'January': 'Janeiro', 'February': 'Fevereiro', 'March': 'Março',
                    'April': 'Abril', 'May': 'Maio', 'June': 'Junho',
                    'July': 'Julho', 'August': 'Agosto', 'September': 'Setembro',
                    'October': 'Outubro', 'November': 'Novembro', 'December': 'Dezembro'
                }
                for en, pt in meses_pt.items():
                    mes_nome = mes_nome.replace(en, pt)
                
                promo_mes['titulo'] = f"{titulo_base} - {mes_nome}"
                
                promocoes_mensais.append(promo_mes)
                
                # Avança para o próximo mês
                data_atual = proximo_mes
            
            logger.info(f"📅 Promoção dividida em {len(promocoes_mensais)} meses")
            return promocoes_mensais
            
        except Exception as e:
            logger.warning(f"⚠️ Erro ao dividir por meses: {e}. Usando promoção original.")
            return [promo_data]
    
    def generate_promotion_excel(self, promo_data: Dict) -> str:
        """
        Gera um arquivo Excel com os dados da promoção (formato horizontal - colunas)
        Se o período for > 1 mês, cria múltiplas linhas automaticamente
        
        Args:
            promo_data: Dicionário com dados da promoção
            
        Returns:
            str: Caminho do arquivo gerado
        """
        try:
            # Divide por meses se necessário
            promocoes = self._dividir_por_meses(promo_data)
            
            # Se resultou em múltiplas promoções, usa o método de múltiplas
            if len(promocoes) > 1:
                return self.generate_multiple_promotions_excel(promocoes)
            
            # Caso contrário, gera normalmente (1 linha)
            # Cria workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Promoção"
            
            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            data_font = Font(size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Extrai todos os percentuais de desconto (para progressivas)
            def extrair_percentuais(recompensas_text, mecanica):
                """Extrai todos os percentuais de uma string de recompensas"""
                if not recompensas_text or mecanica != 'progressiva':
                    return promo_data.get('desconto_percentual', '')
                
                import re
                # Procura por padrões como "5% OFF", "7%", etc.
                percentuais = re.findall(r'(\d+(?:\.\d+)?)\s*%', str(recompensas_text))
                if percentuais:
                    return "; ".join(f"{p}%" for p in percentuais)
                return promo_data.get('desconto_percentual', '')
            
            descontos = extrair_percentuais(
                promo_data.get('recompensas', ''),
                promo_data.get('mecanica', '')
            )
            
            # Define colunas e seus dados (formato horizontal)
            columns = [
                ("Título", promo_data.get('titulo', '')),
                ("Mecânica", promo_data.get('mecanica', '')),
                ("Descrição", promo_data.get('descricao', '')),
                ("Público-Alvo", promo_data.get('segmentacao', '')),
                ("Canal", promo_data.get('canal', '')),
                ("Cluster", promo_data.get('cluster', '')),
                ("Data Início", promo_data.get('periodo_inicio', '')),
                ("Data Fim", promo_data.get('periodo_fim', '')),
                ("Condições", promo_data.get('condicoes', '')),
                ("Recompensas", promo_data.get('recompensas', '')),
                ("Desconto %", descontos),
                ("Produtos", promo_data.get('produtos', '')),
                ("Categoria", promo_data.get('categoria', '')),
                ("Grupo", promo_data.get('grupo', '')),
                ("Qt. Mínima", promo_data.get('qt_minima', '')),
                ("Volume Mínimo", promo_data.get('volume_minimo', '')),
                ("Fabricante", promo_data.get('fabricante', '')),
                ("Código Interno", promo_data.get('codigo_interno', '')),
                ("Observações", promo_data.get('observacoes', '')),
            ]
            
            # Linha 1: Cabeçalhos (azul)
            for col_num, (header, _) in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Linha 2: Dados
            for col_num, (_, value) in enumerate(columns, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.value = str(value) if value else ""
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Ajusta altura das linhas
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 60
            
            # Ajusta largura das colunas
            for col_num in range(1, len(columns) + 1):
                column_letter = get_column_letter(col_num)
                # Larguras específicas para colunas importantes
                if col_num in [1, 3, 9, 10]:  # Título, Descrição, Condições, Recompensas
                    ws.column_dimensions[column_letter].width = 35
                elif col_num in [4, 5]:  # Público, Canal
                    ws.column_dimensions[column_letter].width = 20
                else:
                    ws.column_dimensions[column_letter].width = 15
            
            # Informações de geração (abaixo dos dados)
            row = 4
            ws[f'A{row}'] = f"📅 Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            ws[f'A{row}'].font = Font(italic=True, size=9)
            
            # Nome do arquivo
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            titulo_safe = promo_data.get('titulo', 'promocao').replace(' ', '_')[:30]
            filename = f"promocao_{titulo_safe}_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            
            # Salva arquivo
            wb.save(filepath)
            logger.info(f"✅ Arquivo Excel gerado: {filepath}")
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"❌ Erro ao gerar Excel: {e}", exc_info=True)
            raise
    
    def generate_multiple_promotions_excel(self, promotions: list) -> str:
        """
        Gera um arquivo Excel com múltiplas promoções
        
        Args:
            promotions: Lista de dicionários com dados de promoções
            
        Returns:
            str: Caminho do arquivo gerado
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Promoções"
            
            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Cabeçalhos
            headers = [
                "Título", "Tipo", "Descrição", "Público-Alvo", 
                "Início", "Fim", "Condições", "Recompensas"
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Dados
            for row_num, promo in enumerate(promotions, 2):
                ws.cell(row=row_num, column=1).value = promo.get('titulo', '')
                ws.cell(row=row_num, column=2).value = promo.get('mecanica', '')
                ws.cell(row=row_num, column=3).value = promo.get('descricao', '')
                ws.cell(row=row_num, column=4).value = promo.get('segmentacao', '')
                ws.cell(row=row_num, column=5).value = promo.get('periodo_inicio', '')
                ws.cell(row=row_num, column=6).value = promo.get('periodo_fim', '')
                ws.cell(row=row_num, column=7).value = promo.get('condicoes', '')
                ws.cell(row=row_num, column=8).value = promo.get('recompensas', '')
                
                # Aplica bordas
                for col_num in range(1, 9):
                    ws.cell(row=row_num, column=col_num).border = border
            
            # Ajusta largura das colunas
            for col_num in range(1, 9):
                ws.column_dimensions[get_column_letter(col_num)].width = 20
            
            # Nome do arquivo
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"promocoes_lista_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            
            # Salva arquivo
            wb.save(filepath)
            logger.info(f"✅ Arquivo Excel com {len(promotions)} promoções gerado: {filepath}")
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"❌ Erro ao gerar Excel: {e}", exc_info=True)
            raise


# Instância global
excel_service = ExcelService()


def gerar_excel_promocao(promo_data: Dict) -> str:
    """
    Função auxiliar para gerar Excel de uma promoção
    
    Args:
        promo_data: Dados da promoção
        
    Returns:
        str: Caminho do arquivo gerado
    """
    return excel_service.generate_promotion_excel(promo_data)
